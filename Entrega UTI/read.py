from os import listdir

import db
import dict
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


def ReadInput(path_input, path_db):
    files = []
    onlyfiles = [f for f in listdir(path_input)]
    rows = db.Select(path_db)

    for file in onlyfiles:
        if file.endswith(".xlsx"):
            if file not in rows:
                files.append(file)
            else:
                print("Archivo: " + file + " ya fue procesado con anterioridad.")
    return files


def ReadXLSX(path, file):
    full_file = path + file  # DIRECTORIO DE ARCHIVO INPUT
    values_datos_generales = {}  # DECLARACION DE DICCIONARIO DE DATOS GENERALES
    values_monitoreo = {}  # DECLARACION DE DICCIONARIO DE MONITOREO
    wb = load_workbook(full_file, data_only=True)  # ABRE LIBRO XLSX

    # sh_general = 'General'
    sh_monitoreo = 'Monitoreo'  # NOMBRE DE PESATAÑA
    sh_estadistica = 'Estadística nueva'  # NOMBRE DE PESATAÑA
    sh = getSheet(sh_estadistica, wb)  # ABRE PESTAÑA ESTADISTICA
    try:
        if sh is not None:
            rows, fin_columns = getEstadistica_Context()  # CONSULTA EN DICCIONARIO ESTADISTICA_CONTEXT_INPUT DESDE/HASTA
            aux = ''
            for row in sh.iter_rows(rows, rows, sh.min_column, fin_columns):  # ITERACION DESDE HASTA
                for cell in row:
                    context_generic = {}
                    context_range_title = str(cell.column_letter) + str(cell.row - 8)
                    context_range = str(cell.column_letter) + str(cell.row - 7)
                    context_title = str(sh[context_range_title].value)
                    col_title = str(sh[context_range].value)

                    if sh[context_range_title].value is None:
                        context_title = aux
                    else:
                        aux = context_title

                    context_name = changeContext(context_title, col_title)

                    value = str(cell.value)
                    col_destino = str(cell.column_letter)
                    context_generic['value'] = value
                    context_generic['col_destino'] = col_destino

                    if context_name not in values_datos_generales:
                        values_datos_generales[context_name] = context_generic
                    else:
                        context_name = context_name + "_2"
                        values_datos_generales[context_name] = context_generic

            # HOJA MONITOREO
            sh = getSheet(sh_monitoreo, wb)  # MONITOREO
            if sh is not None:
                for option in range(0, 5):
                    result = getContext(option)
                    context = result[0]
                    custom_range = result[1]
                    context_generic = findValues(custom_range, sh)
                    values_monitoreo[context] = context_generic
                print("Archivo: " + file + " cargado.")

            else:
                print('Imposible identificar hoja ' + sh_monitoreo + ' en libro: ' + file)
                values_monitoreo.clear()
                values_datos_generales.clear()
                values_datos_generales['file'] = file
        else:
            print('Imposible identificar hoja ' + sh_estadistica + ' en libro: ' + file)
            values_monitoreo.clear()
            values_datos_generales.clear()
            values_monitoreo['file'] = file

        file_dict = {  # GUARDA DATOS DEL ARCHIVO PROCESADO
            'value': file,
            'col_destino': 'ANC'
        }

        values_datos_generales['file'] = file_dict

    except NameError:
        print('Error en archivo ' + file)

    wb.close()
    return values_datos_generales, values_monitoreo


def findValues(custom_range, sh):
    context_generic = {}
    for row in custom_range:
        col_inicio, col_fin = getMononitoreo_Context()
        result = getMonitoreo(str(row))
        key = result[0]
        find_row = result[1]
        lst_generic = getValues(col_inicio, col_fin, find_row, sh)
        context_generic[key] = lst_generic
    return context_generic


def getValues(col_i, col_f, row, sh):
    try:
        lst = []
        for col in range(col_i, col_f):
            col_letter = get_column_letter(col)
            x = sh[str(col_letter) + str(row)].value
            if x is None:
                x = ''
            lst.append(x)
    except NameError:
        print('Error')
        return 'Error'
    return lst


def setValues(cell, sh):
    value = sh[cell].value
    if value is None:
        value = ''
    return value


def getDatosGenerales(option):
    return dict.datos_generales_input.get(option, 9)


def getMonitoreo(row):
    return dict.monitoreo_input.get(row, 9)


def getContext(option):
    return dict.contexto_input.get(option, 9)


def getEstadistica_Context():
    rows = dict.estadistica_contexto_input.get('rows')
    columns = dict.estadistica_contexto_input.get('columns')
    return rows, columns


def getMononitoreo_Context():
    col_i = dict.monitoreo_contexto_input.get('col_inicio')
    col_f = dict.monitoreo_contexto_input.get('col_fin')
    return col_i, col_f


def getSheet(sheet_name, wb):
    sheets = wb.get_sheet_names()
    if sheet_name in sheets:
        index = sheets.index(sheet_name)
        return wb[sheets[index]]
    else:
        print('No existe Hoja: ' + sheet_name)
        return None


def changeContext(context_title, col_title):
    new_context_title = ''.join(char for char in context_title if char.isalnum())
    new_col_title = ''.join(char for char in col_title if char.isalnum())
    return new_context_title + "_" + new_col_title
