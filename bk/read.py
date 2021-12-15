import os
from os import listdir
import dict
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from openpyxl.worksheet import worksheet
from openpyxl.utils import get_column_letter


def ReadInput(path_input):
    files = []
    onlyfiles = [f for f in listdir(path_input)]
    for file in onlyfiles:
        if file.endswith(".xlsx"):
            files.append(file)
    return files


def ReadXLSX(path, file):
    full_file = path + file
    values = {}
    wb = load_workbook(full_file, data_only=True)
    # HOJA GENERAL
    sh_general = 'General'
    sh_monitoreo = 'Monitoreo'
    sh = getSheet(sh_general, wb)

    if sh is not None:
        dic_range = len(dict.datos_generales_input)
        for option in range(0, dic_range):
            result = getDatosGenerales(option)
            context = result[0]
            cell = result[1]
            value = setValues(cell, sh)
            values[context] = value

        # HOJA MONITOREO
        sh = getSheet(sh_monitoreo, wb)  # MONITOREO
        if sh is not None:
            for option in range(0, 5):
                result = getContext(option)
                context = result[0]
                custom_range = result[1]
                context_generic = findValues(custom_range, sh)
                values[context] = context_generic
        else:
            print('Imposible identificar hoja Monitoreo en libro: ' + file)
            values.clear()
            values['file'] = file
    else:
        print('Imposible identificar hoja General en libro: ' + file)
        values.clear()
        values['file'] = file
    wb.close()
    return values


def findValues(custom_range, sh):
    context_generic = {}
    for row in custom_range:
        col_inicio = 3
        col_fin = 30
        result = getMonitoreo(str(row))
        key = result[0]
        find_row = result[1]
        lst_generic = getValues(col_inicio, col_fin, find_row, sh)
        context_generic[key] = lst_generic
    return context_generic


def getValues(col_i, col_f, row, sh):
    try:
        lst = []
        for col in range(col_i, col_f):
            col_letter = get_column_letter(col)
            x = sh[str(col_letter) + str(row)].value
            if x is None:
                x = ''
            lst.append(x)
    except NameError:
        print('Error')
        return 'Error'
    return lst


def setValues(cell, sh):
    value = sh[cell].value
    if value is None:
        value = ''
    return value


def getDatosGenerales(option):
    return dict.datos_generales_input.get(option, 9)


def getMonitoreo(row):
    return dict.monitoreo_input.get(row, 9)


def getContext(option):
    return dict.contexto_input.get(option, 9)


def getSheet(sheet_name, wb):
    sheets = wb.get_sheet_names()
    if sheet_name in sheets:
        index = sheets.index(sheet_name)
        return wb[sheets[index]]
    else:
        print('No existe Hoja: ' + sheet_name)
        return None
