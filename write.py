import os

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import column_index_from_string
from datetime import datetime
import db
import dict
import read


def WriteXLSX(path, file, lst, path_output, path_db):
    full_file = path + file
    sheet_template = 'Estadística nueva'
    file_output = path_output + 'Output.xlsx'
    if Exist_File(file_output):
        full_file = file_output
    else:
        full_file = full_file
    try:
        wb = load_workbook(full_file)
        sh = read.getSheet(sheet_template, wb)
        if sh is not None:
            if full_file == file_output:
                row_inicio, columna_destino = getEstadistica_Context_Output()
                row_inicio = sh.max_row + 1  # OBTENER EL ÚLTIMO REGISTRO
            else:
                row_inicio, columna_destino = getEstadistica_Context_Output()
            for values in lst:
                estadisticas = values[0]
                monitoreos = values[1]
                for e in estadisticas:
                    columna_destino = estadisticas[e]['col_destino']
                    sh[columna_destino + str(row_inicio)].value = estadisticas[e]['value']
                for m in monitoreos:
                    context = monitoreos[m]
                    for c in context:
                        columna_destino = getColumnMonitoreo(c)
                        if columna_destino is not 'None':
                            for v in context[c]:
                                sh[columna_destino + str(row_inicio)].value = v
                                index = column_index_from_string(columna_destino)
                                index += 1
                                columna_destino = get_column_letter(index)
                        else:
                            pass

                row_inicio += 1
        else:
            print('Imposible identificar hoja Monitoreo en libro: ' + file)
        wb.save(file_output)
        for values in lst:
            estadisticas = values[0]
            f = estadisticas['file']['value']
            print("Archivo: " + f + " procesado.")
            now = datetime.now()
            dt = now.strftime("%d/%m/%Y %H:%M:%S")
            db.InsertDB(path_db, f, dt)
    except NameError:
        print('Error!')
    return file_output


def getEstadistica_Context_Output():
    rows = dict.estadistica_contexto_output.get('row_inicio')
    columns = dict.estadistica_contexto_output.get('col_fin')
    return rows, columns


def getMonitoreo_Context_Output():
    col_inicio = dict.estadistica_contexto_output.get('col_inicio')
    return col_inicio


def getColumnMonitoreo(key):
    col = dict.monitoreo_output.get(key)
    return col[0]


def Exist_File(file_output):
    if os.path.isfile(file_output):
        return True
    else:
        return False
